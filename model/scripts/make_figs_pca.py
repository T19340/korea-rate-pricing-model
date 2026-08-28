# -*- coding: utf-8 -*-
"""PCA 병행 재추정 Exhibit — 리포트용 SVG 2종(JPM-tone) + Word용 논문식 1종.

입력: rawdata/PCA/BOK_PCA_1D_1W_1M_Shock_Result_평균회귀X.xlsx (부서 병행 분석),
      output/m4_betas.csv · m4_scenarios.csv (대조용 본 리포트 추정)
산출: figs/ex_pca_structure.svg — (a) 요인 1점당 만기별 반응 (b) 당일 프로파일 3잣대 대조
      figs/ex_pca_horizon.svg   — 1D/1W/1M 시간 전개 + 1M 95% 부트스트랩 대역
      figs_paper/fig13_pca_shock.png/.svg — 논문식 2패널판 (Word 그림 7)
"""
from __future__ import annotations

import csv

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl

from common import FIGS, OUT, RAW

FIGP = OUT.parents[0] / "figs_paper"
NAVY, GOLD, TEAL = "#16487E", "#8F7332", "#00838F"
GRID, AXIS, MUT = "#D6D6D6", "#0E2A47", "#6E6E6E"
BLACK, G3, G5, G7 = "#000000", "#4d4d4d", "#808080", "#b8b8b8"

JPM_RC = {
    "svg.fonttype": "none", "font.family": "sans-serif",
    "font.sans-serif": ["Malgun Gothic", "Inter", "Segoe UI", "sans-serif"],
    "axes.unicode_minus": False, "font.size": 10.5,
    "axes.edgecolor": AXIS, "axes.linewidth": 1.2,
    "axes.grid": True, "grid.color": GRID, "grid.linewidth": 0.5,
    "axes.axisbelow": True, "xtick.color": MUT, "ytick.color": MUT,
    "axes.labelcolor": MUT,
    "figure.facecolor": "white", "axes.facecolor": "white",
}
PAPER_RC = {
    "svg.fonttype": "none", "font.family": "serif",
    "font.serif": ["Batang", "Times New Roman", "DejaVu Serif"],
    "axes.unicode_minus": False, "font.size": 9, "axes.labelsize": 9,
    "axes.edgecolor": "black", "axes.linewidth": 0.7,
    "xtick.direction": "in", "ytick.direction": "in",
    "xtick.top": True, "ytick.right": True,
    "xtick.major.size": 3.0, "ytick.major.size": 3.0,
    "xtick.labelsize": 8, "ytick.labelsize": 8, "axes.grid": False,
    "legend.fontsize": 8, "legend.frameon": True, "legend.fancybox": False,
    "legend.edgecolor": "black", "legend.framealpha": 1.0,
    "figure.facecolor": "white", "axes.facecolor": "white",
    "lines.linewidth": 1.1,
}

# ---------------- 데이터
XL = RAW / "PCA" / "BOK_PCA_1D_1W_1M_Shock_Result_평균회귀X.xlsx"
wb = openpyxl.load_workbook(XL, read_only=True, data_only=True)

ws = wb["Scenario_Result"]
hdr = [str(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
col = {h: i for i, h in enumerate(hdr)}
yrs, e1d, e1w, e1m, lo95, hi95 = [], [], [], [], [], []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    yrs.append(float(row[col["Maturity_Years"]]))
    e1d.append(float(row[col["Expected_1D_Shock_bp"]]))
    e1w.append(float(row[col["Expected_1W_Shock_bp"]]))
    e1m.append(float(row[col["Expected_1M_Shock_bp"]]))
    lo95.append(float(row[col["1M_Low95_bp"]]))
    hi95.append(float(row[col["1M_High95_bp"]]))

ws = wb["Factor_Shock_1Score"]
f_yrs_lbl, f_lv, f_sl, f_cv = [], [], [], []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    f_yrs_lbl.append(str(row[0]))
    f_lv.append(float(row[1])); f_sl.append(float(row[2])); f_cv.append(float(row[3]))
wb.close()
f_yrs = list(yrs)  # 같은 15개 만기 격자

yrs = np.array(yrs)
pca_norm = np.array(e1d) / e1d[0]

with open(OUT / "m4_betas.csv", encoding="utf-8-sig") as f:
    b_rows = list(csv.DictReader(f))
b_tau = np.array([float(r["tau_yr"]) for r in b_rows])
b_val = np.array([float(r["beta"]) for r in b_rows])
with open(OUT / "m4_scenarios.csv", encoding="utf-8-sig") as f:
    s_rows = list(csv.DictReader(f))
d_tau = np.array([float(r["tau_yr"]) for r in s_rows])
d_val = np.array([float(r["dns_slope_loading_norm"]) for r in s_rows])

TICKS = [0.25, 0.5, 1, 2, 3, 5, 10, 20, 30]
TLBL = ["3M", "6M", "1Y", "2Y", "3Y", "5Y", "10Y", "20Y", "30Y"]


def maturity_axis(ax):
    ax.set_xscale("log")
    ax.set_xticks(TICKS)
    ax.set_xticklabels(TLBL)
    ax.minorticks_off()
    ax.set_xlim(0.22, 34)


def despine(ax, keep=("left", "bottom")):
    for side in ("top", "right", "left", "bottom"):
        ax.spines[side].set_visible(side in keep)


# ---------------- 리포트용 (JPM-tone)
with plt.rc_context(JPM_RC):
    # EX: 구조 — (a) 요인 1점당 반응 (b) 당일 프로파일 대조
    fig, (a, b) = plt.subplots(1, 2, figsize=(10.6, 4.0))
    a.axhline(0, color="black", lw=1.2, zorder=1)
    a.plot(f_yrs, f_lv, color=NAVY, lw=2.2, marker="o", ms=4, label="수준")
    a.plot(f_yrs, f_sl, color=GOLD, lw=1.6, ls="--", marker="s", ms=4,
           markerfacecolor="white", label="기울기")
    a.plot(f_yrs, f_cv, color=TEAL, lw=1.4, ls=":", marker="^", ms=4,
           markerfacecolor="white", label="곡률")
    maturity_axis(a); despine(a)
    a.set_ylabel("요인 1점당 반응 (bp)")
    a.set_title("(a) PCA 요인의 만기별 전파(적재×척도)", fontsize=10.5,
                color=AXIS, loc="left")
    a.legend(frameon=False, loc="upper left", fontsize=9)

    b.plot(b_tau, b_val, color=NAVY, lw=2.2, marker="o", ms=4.5, label="실증 베타")
    b.plot(d_tau, d_val, color=GOLD, lw=1.6, ls="--", marker="s", ms=4.5,
           markerfacecolor="white", label="DNS 전파")
    b.plot(yrs, pca_norm, color=TEAL, lw=2.0, ls="-.", marker="^", ms=4.5,
           markerfacecolor="white", label="PCA 재추정")
    maturity_axis(b); despine(b)
    b.set_xlim(0.22, 34)
    b.set_ylim(0, 1.08)
    b.set_ylabel("당일 반응 프로파일 (3개월물=1)")
    b.set_title("(b) 당일 프로파일 — 세 잣대의 대조", fontsize=10.5,
                color=AXIS, loc="left")
    b.legend(frameon=False, loc="upper right", fontsize=9)
    fig.tight_layout()
    fig.savefig(FIGS / "ex_pca_structure.svg", bbox_inches="tight", pad_inches=0.15)
    plt.close(fig)
    print("saved ex_pca_structure.svg")

    # EX: 시간 전개 — 1D/1W/1M + 1M 95% 대역
    fig, ax = plt.subplots(figsize=(9.2, 4.4))
    ax.fill_between(yrs, lo95, hi95, color=TEAL, alpha=0.10, lw=0,
                    label="1개월 95% 부트스트랩 구간")
    ax.axhline(0, color="black", lw=1.2, zorder=1)
    ax.plot(yrs, e1d, color=NAVY, lw=2.2, marker="o", ms=4.5, label="당일 (1D)")
    ax.plot(yrs, e1w, color=GOLD, lw=1.6, ls="--", marker="s", ms=4.5,
            markerfacecolor="white", label="1주 (1W)")
    ax.plot(yrs, e1m, color=TEAL, lw=2.4, marker="^", ms=5, label="1개월 (1M)")
    maturity_axis(ax); despine(ax)
    ax.set_ylabel("+25bp 컨센서스 서프라이즈 시 반응 (bp)")
    ax.annotate("당일: 3M에 집중", xy=(0.25, e1d[0]), xytext=(0.42, 12.6),
                color=NAVY, fontsize=9.5,
                arrowprops=dict(arrowstyle="-", color=NAVY, lw=0.8))
    ax.annotate("1개월: 장기로 확산", xy=(10, e1m[11]), xytext=(3.3, 13.6),
                color=TEAL, fontsize=9.5,
                arrowprops=dict(arrowstyle="-", color=TEAL, lw=0.8))
    ax.legend(frameon=False, loc="lower right", fontsize=9)
    fig.tight_layout()
    fig.savefig(FIGS / "ex_pca_horizon.svg", bbox_inches="tight", pad_inches=0.15)
    plt.close(fig)
    print("saved ex_pca_horizon.svg")

# ---------------- Word용 (논문식)
with plt.rc_context(PAPER_RC):
    fig, (a, b) = plt.subplots(1, 2, figsize=(7.4, 2.9))
    a.plot(b_tau, b_val, color=BLACK, lw=1.1, marker="o", ms=3.5, label="실증 베타")
    a.plot(d_tau, d_val, color=G3, lw=1.1, ls="--", marker="s", ms=3.5,
           markerfacecolor="white", label="DNS 전파")
    a.plot(yrs, pca_norm, color=G5, lw=1.1, ls="-.", marker="^", ms=3.8,
           markerfacecolor="white", label="PCA 재추정")
    maturity_axis(a)
    a.set_ylim(0, 1.08)
    a.set_ylabel("당일 반응 (3개월물=1)")
    a.set_xlabel("만기")
    a.text(0.0, 1.03, "(a)", transform=a.transAxes, fontsize=9)
    leg = a.legend(loc="upper right"); leg.get_frame().set_linewidth(0.6)

    b.fill_between(yrs, lo95, hi95, color=G7, alpha=0.45, lw=0)
    b.axhline(0, color="black", lw=0.7)
    b.plot(yrs, e1d, color=BLACK, lw=1.1, marker="o", ms=3.5, label="당일")
    b.plot(yrs, e1w, color=G3, lw=1.1, ls="--", marker="s", ms=3.5,
           markerfacecolor="white", label="1주")
    b.plot(yrs, e1m, color=G5, lw=1.3, ls="-.", marker="^", ms=3.8, label="1개월")
    maturity_axis(b)
    b.set_ylabel("+25bp 서프라이즈 시 반응 (bp)")
    b.set_xlabel("만기")
    b.text(0.0, 1.03, "(b)", transform=b.transAxes, fontsize=9)
    leg = b.legend(loc="upper left"); leg.get_frame().set_linewidth(0.6)
    fig.tight_layout()
    for ext, kw in [(".png", {"dpi": 300}), (".svg", {})]:
        fig.savefig(FIGP / ("fig13_pca_shock" + ext), bbox_inches="tight",
                    pad_inches=0.05, **kw)
    plt.close(fig)
    print("saved fig13_pca_shock")
