# -*- coding: utf-8 -*-
"""Shared data loading for the analyst report models."""
from __future__ import annotations

import csv
import json
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]          # korea_rate_pricing_model/
BASE = Path(__file__).resolve().parents[1]          # 이 스크립트가 속한 분석 폴더
RAW = ROOT / "rawdata"
OUT = BASE / "output"
FIGS = BASE / "figs"

# ---------------------------------------------------------------- calendars

def mpc_meetings() -> list[date]:
    with open(RAW / "bok" / "mpc_meeting_calendar.csv", encoding="utf-8-sig") as f:
        return [datetime.strptime(r["meeting_date"], "%Y-%m-%d").date()
                for r in csv.DictReader(f)]


# 2027 schedule is not announced yet; quarterly placeholder nodes carry the
# post-November expectation so it does not contaminate the 2026 meetings.
def model_config() -> dict:
    """config/model.json — 모형이 바라보는 기간 설정."""
    with open(ROOT / "config" / "model.json", encoding="utf-8") as f:
        return json.load(f)


CFG = model_config()
TARGET_YEAR = int(CFG["target_year"])

# 금통위 일정 미공표 구간에 두는 가정 노드. 코드가 아니라 설정에서 온다 —
# 사이클이 바뀔 때마다 스크립트를 고치게 만들면 저장소가 곧 낡는다.
ASSUMED_FUTURE_NODES = [datetime.strptime(d, "%Y-%m-%d").date()
                        for d in CFG["assumed_future_nodes"]]
ASSUMED_2027_NODES = ASSUMED_FUTURE_NODES        # 이전 이름 유지(호환)


def target_meetings() -> list[date]:
    """확률을 열로 기록할 대상 연도의 금통위."""
    return [m for m in mpc_meetings() if m.year == TARGET_YEAR]


def base_rate_history() -> list[tuple[date, float]]:
    with open(RAW / "bok" / "policy_rate_change_history.csv", encoding="utf-8-sig") as f:
        rows = [(datetime.strptime(r["effective_date"], "%Y-%m-%d").date(),
                 float(r["rate_percent"])) for r in csv.DictReader(f)]
    return sorted(rows)


def base_rate_on(d: date, hist=None) -> float:
    hist = hist or base_rate_history()
    rate = hist[0][1]
    for eff, r in hist:
        if eff <= d:
            rate = r
        else:
            break
    return rate

# ---------------------------------------------------------------- daily panels

def load_tidy(path: Path) -> dict[date, float]:
    out = {}
    with open(path, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            if r["value"]:
                out[datetime.strptime(r["date"], "%Y-%m-%d").date()] = float(r["value"])
    return out


def ecos_daily(slug: str) -> dict[date, float]:
    return load_tidy(RAW / "ecos_daily" / "rates" / f"{slug}.csv")


def kofr_daily() -> dict[date, float]:
    """KOFR: official-API file, overlaid with the fresher internal-endpoint panel."""
    out = {}
    with open(RAW / "ecos" / "rates" / "kofr_daily.csv", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            if r.get("DATA_VALUE"):
                d = datetime.strptime(r["TIME"], "%Y%m%d").date()
                out[d] = float(r["DATA_VALUE"])
    with open(RAW / "ecos_daily" / "daily_rates_panel.csv", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            if r.get("kofr"):
                out[datetime.strptime(r["date"], "%Y-%m-%d").date()] = float(r["kofr"])
    return out


def add_months(d: date, months: int) -> date:
    month = d.month - 1 + months
    year = d.year + month // 12
    month = month % 12 + 1
    day = min(d.day, [31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)
                      else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month - 1])
    return date(year, month, day)


def tenor_days(asof: date, tenor_kr: str) -> int:
    if tenor_kr == "1주":
        return 7
    n = {"1개월": 1, "2개월": 2, "3개월": 3, "6개월": 6, "9개월": 9,
         "1년": 12, "18개월": 18, "2년": 24, "3년": 36}[tenor_kr]
    return (add_months(asof, n) - asof).days


def kofia_avg(slug: str) -> dict[date, float]:
    out = {}
    with open(RAW / "kofia" / "valuation" / f"{slug}.csv", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            v = r.get("evaluator_avg", "")
            if v:
                out[datetime.strptime(r["date"], "%Y-%m-%d").date()] = float(v)
    return out

# ---------------------------------------------------------------- infomax files

OIS_TENOR_DAYS = {"1주": 7, "1개월": 30, "2개월": 61, "3개월": 91, "6개월": 182,
                  "9개월": 273, "1년": 365, "18개월": 548, "2년": 730, "3년": 1095}


def load_ois(mid_only: bool = True) -> dict[date, dict[str, float]]:
    """KOFR OIS quotes: {date: {tenor_kr: mid}} for tenors up to 3Y."""
    wb = openpyxl.load_workbook(RAW / "infomax" / "KOFR_OIS.xlsx", read_only=True)
    ws = wb["Sheet1"]
    import re
    r2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    titles = []
    for c in r2:
        if c:
            m = re.search(r"한국자금 ([^,]+),", str(c))
            titles.append(m.group(1) if m else "?")
    out: dict[date, dict[str, float]] = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or row[0] is None:
            continue
        d = row[0].date() if hasattr(row[0], "date") else None
        if d is None:
            continue
        quotes = {}
        for k, t in enumerate(titles):
            if t not in OIS_TENOR_DAYS:
                continue
            col = 1 + 4 * k + 2  # MID
            if col < len(row) and row[col] is not None:
                quotes[t] = float(row[col])
        if quotes:
            out[d] = quotes
    wb.close()
    return out


IRS_KEEP = ["03M", "06M", "09M", "01Y", "18M", "02Y", "03Y", "05Y", "10Y"]


def load_irs() -> dict[date, dict[str, float]]:
    wb = openpyxl.load_workbook(RAW / "infomax" / "IRS_KRWKRW.xlsx", read_only=True)
    ws = wb["Sheet1"]
    header = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    idx = {h: i for i, h in enumerate(header) if h}
    out: dict[date, dict[str, float]] = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or row[0] is None:
            continue
        d = row[0].date() if hasattr(row[0], "date") else None
        if d is None:
            continue
        quotes = {t: float(row[idx[t]]) for t in IRS_KEEP
                  if t in idx and idx[t] < len(row) and row[idx[t]] is not None}
        if quotes:
            out[d] = quotes
    wb.close()
    return out


def business_days(series: dict[date, float], start: date, end: date) -> list[date]:
    return sorted(d for d in series if start <= d <= end)
