# -*- coding: utf-8 -*-
"""m12 — 국면별 당일 프로파일(3개월=1): 실증(통안91 자) vs PCA(컨센서스 자) vs DNS.

실증·PCA 프로파일은 우리 원자료 9개 만기의 발표일 변화로 국면별 원점 회귀
(2013년 이후 — 6개월물 민평 가용 시점). DNS는 기울기 적재함수(국면 불변).
PCA 잣대는 |컨센서스 오차| 5bp 이상 이벤트가 완만기에만 충분(13회)하고,
급속 인상기(4회)는 부호만, 완화·재인상기(3회)는 계산 불가 — 그대로 공시한다.

산출: output/m12_regime_profiles.csv
      figs/ex_regime_profiles.svg (리포트 Exhibit 12)
      figs_paper/fig14_regime_profiles.png/.svg (Word 그림 8)
"""
from __future__ import annotations

import csv
from datetime import date, timedelta

import numpy as np
import openpyxl

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from common import FIGS, OUT, RAW, ecos_daily, kofia_avg, mpc_meetings

FIGP = OUT.parents[0] / "figs_paper"
NAVY, GOLD, TEAL, MUT = "#16487E", "#8F7332", "#00838F", "#6E6E6E"
BLACK, G3, G5 = "#000000", "#4d4d4d", "#808080"

JPM_RC = {
    "svg.fonttype": "none", "font.family": "sans-serif",
    "font.sans-serif": ["Malgun Gothic", "Inter", "Segoe UI", "sans-serif"],
    "axes.unicode_minus": False, "font.size": 10.5,
    "axes.edgecolor": "#0E2A47", "axes.linewidth": 1.2,
    "axes.grid": True, "grid.color": "#D6D6D6", "grid.linewidth": 0.5,
    "axes.axisbelow": True, "xtick.color": MUT, "ytick.color": MUT,
    "axes.labelcolor": MUT,
    "figure.facecolor": "white", "axes.facecolor": "white",
}
PAPER_RC = {
    "svg.fonttype": "none", "font.family": "serif",
    "font.serif": ["Batang", "Times New Roman", "DejaVu Serif"],
    "axes.unicode_minus": False, "font.size": 9, "axes.labelsize": 9,
    "axes.edgecolor": "black", "axes.linewidth": 0.7,
    "xtick.direction": "in", "ytick.direction": "in",
    "xtick.top": True, "ytick.right": True,
    "xtick.major.size": 3.0, "ytick.major.size": 3.0,
    "xtick.labelsize": 7.5, "ytick.labelsize": 8, "axes.grid": False,
    "legend.fontsize": 7.5, "legend.frameon": True, "legend.fancybox": False,
    "legend.edgecolor": "black", "legend.framealpha": 1.0,
    "figure.facecolor": "white", "axes.facecolor": "white",
    "lines.linewidth": 1.1,
}

MATS = [("msb_91d", 0.25), ("msb_6m", 0.5), ("msb_1y", 1.0), ("ktb_2y", 2.0),
        ("ktb_3y", 3.0), ("ktb_5y", 5.0), ("ktb_10y", 10.0),
        ("ktb_20y", 20.0), ("ktb_30y", 30.0)]
TAU = [t for _, t in MATS]
TLBL = ["3M", "6M", "1Y", "2Y", "3Y", "5Y", "10Y", "20Y", "30Y"]

series = {}
for slug, _ in MATS:
    if slug == "msb_6m":
        series[slug] = kofia_avg("msb_6m_val_daily")
    elif slug == "ktb_2y":
        m = dict(kofia_avg("ktb_2y_val_daily"))
        m.update(ecos_daily("ktb_2y_daily"))
        series[slug] = m
    else:
        series[slug] = ecos_daily(f"{slug}_daily")


def day_change(s, d):
    for lag in range(1, 6):
        p = d - timedelta(days=lag)
        if p in s:
            return (s[d] - s[p]) * 100 if d in s else None
    return None


def event_moves(d):
    mv = [day_change(series[s], d) for s, _ in MATS]
    return None if any(v is None for v in mv) else mv


# 상한은 실제 데이터가 끝나는 날까지 (하드코딩하면 표본이 굳는다)
_sample_end = max(series["msb_91d"])
meet = [m for m in mpc_meetings() if date(2013, 1, 1) <= m <= _sample_end]
ours = []
for m in meet:
    s = day_change(series["msb_91d"], m)
    mv = event_moves(m)
    if s is not None and abs(s) >= 1.0 and mv is not None:
        ours.append((m, s, mv))

wb = openpyxl.load_workbook(
    RAW / "PCA" / "BOK_PCA_1D_1W_1M_Shock_Result_평균회귀X.xlsx",
    read_only=True, data_only=True)
ws = wb["All_MPC_Events"]
hdr = [str(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
ix = {h: i for i, h in enumerate(hdr)}
cons = []
for row in ws.iter_rows(min_row=2, values_only=True):
    d = row[ix["MPC_Date"]]
    if d is None:
        continue
    d = d.date() if hasattr(d, "date") else d
    s = row[ix["Surprise_bp"]]
    if s is None or abs(float(s)) < 5.0:
        continue
    mv = event_moves(d)
    if mv is not None:
        cons.append((d, float(s), mv))
wb.close()

with open(OUT / "m4_scenarios.csv", encoding="utf-8-sig") as f:
    DNS = [float(r["dns_slope_loading_norm"]) for r in csv.DictReader(f)]


def beta0(x, y):
    x, y = np.asarray(x, float), np.asarray(y, float)
    return float((x * y).sum() / (x * x).sum())


def profile(events):
    if not events:
        return None
    s = np.array([e[1] for e in events])
    mv = np.array([e[2] for e in events])
    b = np.array([beta0(s, mv[:, j]) for j in range(9)])
    return b / b[0]


REG = [("완만기", date(2013, 1, 1), date(2019, 6, 30)),
       ("급속 인상", date(2021, 8, 1), date(2023, 1, 31)),
       ("완화·재인상", date(2024, 10, 1), date(2026, 12, 31))]
panels = []
for name, s0, s1 in REG:
    o = [e for e in ours if s0 <= e[0] <= s1]
    p = [e for e in cons if s0 <= e[0] <= s1]
    panels.append({"name": name, "n_emp": len(o), "emp": profile(o),
                   "n_pca": len(p),
                   "pca": profile(p) if len(p) >= 4 else None})

with open(OUT / "m12_regime_profiles.csv", "w", newline="",
          encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(["maturity", "tau_yr", "dns"] +
               [f"{k}_{p['name']}" for p in panels for k in ("emp", "pca")])
    for j in range(9):
        row = [TLBL[j], TAU[j], round(DNS[j], 3)]
        for p in panels:
            row.append(round(p["emp"][j], 3) if p["emp"] is not None else "")
            row.append(round(p["pca"][j], 3) if p["pca"] is not None else "")
        w.writerow(row)
print("saved m12_regime_profiles.csv")
for p in panels:
    print(f"  {p['name']}: emp n={p['n_emp']}, pca n={p['n_pca']}"
          f"{'' if p['pca'] is not None else ' (측정 불가)'}")

TITLES_J = ["① 완만기 2013~2019.6", "② 급속 인상 2021.8~2023.1",
            "③ 완화·재인상 2024.10~ (현재 국면)"]
WORLD = ["커브가 통째로 움직인 시대", "인상 서프라이즈 = 경기 하강 뉴스",
         "서프라이즈가 사이클 재평가의 방아쇠"]


def draw(rc, jpm):
    with plt.rc_context(rc):
        fig, axes = plt.subplots(1, 3, figsize=(11.4, 4.2) if jpm else (7.6, 2.9),
                                 sharey=True)
        for i, (ax, p) in enumerate(zip(axes, panels)):
            if jpm and i == 2:
                ax.set_facecolor("#F4F8FC")
            ax.axhline(0, color="black", lw=1.0 if jpm else 0.7)
            ax.axhline(1, color=MUT if jpm else G5, lw=0.7, ls=":")
            c_dns = GOLD if jpm else G5
            c_emp = NAVY if jpm else BLACK
            c_pca = TEAL if jpm else G3
            ax.plot(TAU, DNS, color=c_dns, lw=1.6 if jpm else 1.0, ls="--",
                    marker="s", ms=4 if jpm else 3.2, markerfacecolor="white",
                    label="DNS 전파(국면 불변)")
            ax.plot(TAU, p["emp"], color=c_emp, lw=2.2 if jpm else 1.2,
                    marker="o", ms=4.5 if jpm else 3.4,
                    label=f"실증(통안91 자, n={p['n_emp']})")
            if p["pca"] is not None:
                small = p["n_pca"] < 10
                ax.plot(TAU, p["pca"], color=c_pca, lw=2.0 if jpm else 1.1,
                        ls=":" if small else "-.", marker="^",
                        ms=4.5 if jpm else 3.6, markerfacecolor="white",
                        label=f"PCA(컨센서스 자, n={p['n_pca']})"
                              + (" — 부호만" if small else ""))
            ax.set_xscale("log")
            ax.set_xticks(TAU)
            ax.set_xticklabels(TLBL)
            ax.minorticks_off()
            if jpm:
                ax.set_title(TITLES_J[i], fontsize=11, color="#0E2A47")
                ax.text(0.5, 0.035, WORLD[i], transform=ax.transAxes,
                        ha="center", fontsize=8.5, color=MUT,
                        bbox=dict(facecolor=ax.get_facecolor(), alpha=0.9,
                                  edgecolor="none", pad=1.6))
                ax.legend(frameon=False, fontsize=8.5, loc="upper right")
            else:
                ax.set_title(f"({'abc'[i]}) {TITLES_J[i][2:]}", fontsize=8.5)
                ax.set_xlabel("만기")
                leg = ax.legend(loc="upper right")
                leg.get_frame().set_linewidth(0.6)
        axes[0].set_ylabel("당일 반응 프로파일 (3개월물=1)")
        axes[0].set_ylim(-2.5, 3.4)
        if jpm:
            for yy, tt in [(1.0, "3개월과 같은 폭"), (0.0, "무반응")]:
                axes[0].annotate(tt, xy=(0.28, yy + 0.12), fontsize=8,
                                 color=MUT, ha="left", va="bottom",
                                 bbox=dict(facecolor="white", alpha=0.9,
                                           edgecolor="none", pad=1.2))
        fig.tight_layout()
        if jpm:
            fig.savefig(FIGS / "ex_regime_profiles.svg", bbox_inches="tight",
                        pad_inches=0.15)
            print("saved ex_regime_profiles.svg")
        else:
            for ext, kw in [(".png", {"dpi": 300}), (".svg", {})]:
                fig.savefig(FIGP / ("fig14_regime_profiles" + ext),
                            bbox_inches="tight", pad_inches=0.05, **kw)
            print("saved fig14_regime_profiles")
        plt.close(fig)


draw(JPM_RC, True)
draw(PAPER_RC, False)
